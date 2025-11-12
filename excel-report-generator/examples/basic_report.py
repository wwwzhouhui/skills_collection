#!/usr/bin/env python3
"""
基础 Excel 报表生成示例

演示如何从 CSV 文件或 DataFrame 生成基本的 Excel 报表
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def create_basic_report(data_source, output_file=None):
    """
    创建基础 Excel 报表

    参数:
        data_source: str 或 DataFrame - CSV 文件路径或 pandas DataFrame
        output_file: str - 输出文件名（可选，默认使用时间戳）

    返回:
        str - 生成的文件路径
    """
    # Step 1: 加载数据
    if isinstance(data_source, str):
        df = pd.read_csv(data_source)
        print(f"✓ 已加载数据文件: {data_source}")
    elif isinstance(data_source, pd.DataFrame):
        df = data_source
        print("✓ 已加载 DataFrame 数据")
    else:
        raise ValueError("数据源必须是 CSV 文件路径或 pandas DataFrame")

    # Step 2: 数据预处理
    print(f"  数据行数: {len(df)}")
    print(f"  数据列数: {len(df.columns)}")

    # Step 3: 生成输出文件名
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'report_{timestamp}.xlsx'

    # Step 4: 创建 Excel 文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入数据到工作表
        df.to_excel(writer, sheet_name='Data', index=False)

        # 获取 workbook 和 worksheet 对象
        workbook = writer.book
        worksheet = writer.sheets['Data']

        # Step 5: 应用格式
        apply_basic_formatting(worksheet, df)

    print(f"✓ 报表已生成: {output_file}")
    return output_file


def apply_basic_formatting(worksheet, df):
    """应用基础格式到工作表"""

    # 标题行样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用标题行格式
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 应用数据行边框
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为 50
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    worksheet.freeze_panes = 'A2'

    # 添加自动筛选
    worksheet.auto_filter.ref = worksheet.dimensions


def create_summary_sheet(df, writer, sheet_name='Summary'):
    """
    创建数据汇总工作表

    参数:
        df: DataFrame - 原始数据
        writer: ExcelWriter - Excel 写入器
        sheet_name: str - 工作表名称
    """
    # 生成汇总统计
    summary = df.describe()

    # 写入汇总数据
    summary.to_excel(writer, sheet_name=sheet_name)

    # 获取工作表并格式化
    worksheet = writer.sheets[sheet_name]

    # 标题格式
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 第一列格式（统计指标名称）
    for cell in worksheet['A']:
        if cell.row > 1:
            cell.font = Font(bold=True)

    print(f"✓ 已创建汇总工作表: {sheet_name}")


# 示例用法
if __name__ == "__main__":
    # 示例 1: 从 CSV 生成报表
    # create_basic_report('sales_data.csv')

    # 示例 2: 从 DataFrame 生成报表
    sample_data = {
        'Product': ['A', 'B', 'C', 'D', 'E'],
        'Sales': [1000, 1500, 800, 2000, 1200],
        'Profit': [200, 300, 150, 400, 250],
        'Region': ['East', 'West', 'East', 'North', 'South']
    }
    df = pd.DataFrame(sample_data)

    # 生成报表
    output_file = create_basic_report(df, 'sample_report.xlsx')

    # 添加汇总工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        create_summary_sheet(df, writer)

    print("\n示例报表生成完成！")
    print("查看文件: sample_report.xlsx")
