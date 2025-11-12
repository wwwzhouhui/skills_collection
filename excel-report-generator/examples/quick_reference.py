#!/usr/bin/env python3
"""
Excel 快速参考 - 常用代码片段

这个文件包含了 Excel 报表生成中最常用的代码片段，可以直接复制使用
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


# ============================================================================
# 1. 基础操作
# ============================================================================

def snippet_create_workbook():
    """创建新工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "My Sheet"
    wb.save('workbook.xlsx')


def snippet_load_workbook():
    """加载现有工作簿"""
    wb = load_workbook('existing.xlsx')
    ws = wb.active  # 获取活动工作表
    ws = wb['SheetName']  # 按名称获取工作表


def snippet_read_csv_to_excel():
    """CSV 转 Excel"""
    df = pd.read_csv('data.csv')
    df.to_excel('output.xlsx', sheet_name='Data', index=False)


def snippet_read_excel_to_dataframe():
    """Excel 转 DataFrame"""
    df = pd.read_excel('data.xlsx', sheet_name='Sheet1')
    # 或读取多个工作表
    excel_file = pd.ExcelFile('data.xlsx')
    df1 = excel_file.parse('Sheet1')
    df2 = excel_file.parse('Sheet2')


# ============================================================================
# 2. 单元格操作
# ============================================================================

def snippet_write_cell():
    """写入单元格"""
    wb = Workbook()
    ws = wb.active

    # 方式 1: 直接引用
    ws['A1'] = 'Hello'
    ws['B1'] = 123
    ws['C1'] = datetime.now()

    # 方式 2: 使用 cell()
    ws.cell(row=1, column=1, value='Hello')


def snippet_read_cell():
    """读取单元格"""
    wb = load_workbook('data.xlsx')
    ws = wb.active

    # 读取值
    value = ws['A1'].value
    value = ws.cell(row=1, column=1).value


def snippet_write_dataframe():
    """将 DataFrame 写入工作表"""
    df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    wb = Workbook()
    ws = wb.active

    # 方式 1: 使用 pandas
    with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

    # 方式 2: 使用 openpyxl
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


# ============================================================================
# 3. 样式和格式
# ============================================================================

def snippet_font_style():
    """字体样式"""
    ws['A1'].font = Font(
        name='Arial',
        size=12,
        bold=True,
        italic=False,
        color='FF0000'  # 红色
    )


def snippet_fill_color():
    """单元格填充颜色"""
    ws['A1'].fill = PatternFill(
        start_color='FFFF00',  # 黄色
        end_color='FFFF00',
        fill_type='solid'
    )


def snippet_alignment():
    """对齐方式"""
    ws['A1'].alignment = Alignment(
        horizontal='center',  # left, center, right
        vertical='center',    # top, center, bottom
        wrap_text=True        # 自动换行
    )


def snippet_border():
    """边框"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws['A1'].border = thin_border


def snippet_number_format():
    """数字格式"""
    ws['A1'].number_format = '#,##0.00'  # 千分位，两位小数
    ws['A2'].number_format = '0.00%'     # 百分比
    ws['A3'].number_format = 'yyyy-mm-dd'  # 日期格式
    ws['A4'].number_format = '$#,##0.00'   # 货币格式


def snippet_header_style():
    """标题行完整样式"""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:  # 第一行
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


# ============================================================================
# 4. 行列操作
# ============================================================================

def snippet_adjust_column_width():
    """调整列宽"""
    # 固定宽度
    ws.column_dimensions['A'].width = 20

    # 自动调整（基于内容）
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2


def snippet_adjust_row_height():
    """调整行高"""
    ws.row_dimensions[1].height = 25


def snippet_insert_delete_rows():
    """插入/删除行列"""
    ws.insert_rows(1)  # 在第 1 行前插入一行
    ws.insert_cols(1)  # 在第 1 列前插入一列
    ws.delete_rows(5, 3)  # 从第 5 行开始删除 3 行
    ws.delete_cols(2, 1)  # 删除第 2 列


def snippet_merge_cells():
    """合并单元格"""
    ws.merge_cells('A1:D1')  # 合并 A1 到 D1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # 取消合并
    ws.unmerge_cells('A1:D1')


def snippet_freeze_panes():
    """冻结窗格"""
    ws.freeze_panes = 'A2'  # 冻结第一行
    ws.freeze_panes = 'B1'  # 冻结第一列
    ws.freeze_panes = 'B2'  # 冻结第一行和第一列


# ============================================================================
# 5. 图表
# ============================================================================

def snippet_bar_chart():
    """柱状图"""
    chart = BarChart()
    chart.title = "Sales Chart"
    chart.x_axis.title = "Products"
    chart.y_axis.title = "Sales"

    data = Reference(ws, min_col=2, min_row=1, max_row=10)
    categories = Reference(ws, min_col=1, min_row=2, max_row=10)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E5")


def snippet_line_chart():
    """折线图"""
    chart = LineChart()
    chart.title = "Trend Analysis"
    chart.style = 12

    data = Reference(ws, min_col=2, min_row=1, max_row=10, max_col=3)
    chart.add_data(data, titles_from_data=True)

    ws.add_chart(chart, "E5")


def snippet_pie_chart():
    """饼图"""
    chart = PieChart()
    chart.title = "Market Share"

    data = Reference(ws, min_col=2, min_row=2, max_row=5)
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    ws.add_chart(chart, "E5")


# ============================================================================
# 6. 条件格式
# ============================================================================

def snippet_color_scale():
    """色阶条件格式"""
    ws.conditional_formatting.add(
        'B2:B100',
        ColorScaleRule(
            start_type='min', start_color='F8696B',  # 红色
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # 黄色
            end_type='max', end_color='63BE7B'  # 绿色
        )
    )


def snippet_data_bars():
    """数据条"""
    ws.conditional_formatting.add(
        'C2:C100',
        DataBarRule(
            start_type='min',
            end_type='max',
            color="4472C4"
        )
    )


def snippet_cell_highlight():
    """单元格高亮规则"""
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # 小于 0 的值标红
    ws.conditional_formatting.add(
        'B2:B100',
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
    )

    # 大于 1000 的值标绿
    ws.conditional_formatting.add(
        'B2:B100',
        CellIsRule(operator='greaterThan', formula=['1000'], fill=green_fill)
    )


# ============================================================================
# 7. 公式
# ============================================================================

def snippet_formulas():
    """Excel 公式"""
    ws['D2'] = '=SUM(B2:C2)'  # 求和
    ws['D3'] = '=AVERAGE(B2:B10)'  # 平均值
    ws['D4'] = '=MAX(B2:B10)'  # 最大值
    ws['D5'] = '=MIN(B2:B10)'  # 最小值
    ws['D6'] = '=COUNT(B2:B10)'  # 计数
    ws['D7'] = '=IF(B2>100,"High","Low")'  # 条件判断
    ws['D8'] = '=VLOOKUP(A2,Sheet2!A:B,2,FALSE)'  # 查找


# ============================================================================
# 8. 高级功能
# ============================================================================

def snippet_auto_filter():
    """自动筛选"""
    ws.auto_filter.ref = 'A1:D10'  # 对范围启用筛选
    ws.auto_filter.ref = ws.dimensions  # 对整个数据区域启用筛选


def snippet_data_validation():
    """数据验证（下拉列表）"""
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1='"优秀,良好,一般,较差"')
    ws.add_data_validation(dv)
    dv.add('E2:E100')


def snippet_hyperlink():
    """超链接"""
    ws['A1'].hyperlink = "https://www.example.com"
    ws['A1'].value = "Click Here"
    ws['A1'].font = Font(color='0000FF', underline='single')


def snippet_named_range():
    """命名范围"""
    from openpyxl.workbook.defined_name import DefinedName

    # 创建命名范围
    wb.create_named_range('SalesData', ws, 'A1:D10')

    # 在公式中使用
    ws['F1'] = '=SUM(SalesData)'


def snippet_protect_sheet():
    """保护工作表"""
    ws.protection.sheet = True
    ws.protection.password = 'mypassword'

    # 允许特定单元格编辑
    ws['A1'].protection = Protection(locked=False)


# ============================================================================
# 9. 多工作表操作
# ============================================================================

def snippet_multiple_sheets():
    """多工作表操作"""
    wb = Workbook()

    # 创建工作表
    ws1 = wb.active
    ws1.title = "Sales"

    ws2 = wb.create_sheet("Summary")
    ws3 = wb.create_sheet("Charts", 0)  # 插入到第一个位置

    # 复制工作表
    ws4 = wb.copy_worksheet(ws1)
    ws4.title = "Sales Copy"

    # 删除工作表
    wb.remove(ws4)

    # 遍历所有工作表
    for sheet in wb:
        print(sheet.title)


# ============================================================================
# 10. 性能优化
# ============================================================================

def snippet_write_only_mode():
    """只写模式（大数据集）"""
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    # 写入数据（只能追加）
    for row in range(100000):
        ws.append([f'Row {row}', row * 2, row * 3])

    wb.save('large_file.xlsx')


def snippet_read_only_mode():
    """只读模式（快速读取）"""
    wb = load_workbook('large_file.xlsx', read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            print(cell.value)


# ============================================================================
# 11. 完整示例模板
# ============================================================================

def complete_example_template():
    """完整的报表生成模板"""
    # 1. 准备数据
    df = pd.DataFrame({
        'Product': ['A', 'B', 'C', 'D', 'E'],
        'Sales': [1000, 1500, 800, 2000, 1200],
        'Profit': [200, 300, 150, 400, 250]
    })

    # 2. 创建 Excel
    output_file = 'complete_report.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入数据
        df.to_excel(writer, sheet_name='Data', index=False)

        # 获取对象
        workbook = writer.book
        worksheet = writer.sheets['Data']

        # 3. 格式化标题
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # 4. 调整列宽
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

        # 5. 添加图表
        chart = BarChart()
        chart.title = "Sales Analysis"
        data = Reference(worksheet, min_col=2, min_row=1, max_row=6)
        categories = Reference(worksheet, min_col=1, min_row=2, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        worksheet.add_chart(chart, "E2")

        # 6. 条件格式
        worksheet.conditional_formatting.add(
            'B2:B6',
            ColorScaleRule(start_type='min', start_color='F8696B',
                          end_type='max', end_color='63BE7B')
        )

        # 7. 添加合计行
        last_row = len(df) + 2
        worksheet.cell(row=last_row, column=1, value='TOTAL')
        worksheet.cell(row=last_row, column=2, value=f'=SUM(B2:B{last_row-1})')
        worksheet.cell(row=last_row, column=3, value=f'=SUM(C2:C{last_row-1})')

        for col in range(1, 4):
            worksheet.cell(row=last_row, column=col).font = Font(bold=True)

    print(f"✓ 报表已生成: {output_file}")


# 运行完整示例
if __name__ == "__main__":
    print("Excel 快速参考 - 代码片段集合\n")
    print("每个函数都是独立的代码片段，可以直接复制使用\n")

    # 运行完整示例
    complete_example_template()
