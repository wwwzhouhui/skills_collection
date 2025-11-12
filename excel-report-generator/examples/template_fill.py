#!/usr/bin/env python3
"""
Excel 模板填充示例

演示如何使用预定义的 Excel 模板并填充动态数据
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os


def fill_template(template_path, data, output_file=None):
    """
    基于模板填充数据

    参数:
        template_path: str - 模板文件路径
        data: dict - 要填充的数据字典
        output_file: str - 输出文件名

    返回:
        str - 生成的文件路径
    """
    # 检查模板文件是否存在
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    # 加载模板
    wb = load_workbook(template_path)
    print(f"✓ 已加载模板: {template_path}")

    # 填充数据
    fill_data_cells(wb, data)

    # 生成输出文件名
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = os.path.splitext(os.path.basename(template_path))[0]
        output_file = f'{base_name}_filled_{timestamp}.xlsx'

    # 保存文件
    wb.save(output_file)
    print(f"✓ 报表已生成: {output_file}")

    return output_file


def fill_data_cells(workbook, data):
    """
    填充数据到工作簿中的命名单元格或指定位置

    数据格式示例:
    data = {
        'Sheet1': {
            'A1': 'Company Name',
            'B5': 12345,
            'named_range': 'Value'
        },
        'Sheet2': {
            'C10': datetime.now()
        }
    }
    """
    for sheet_name, sheet_data in data.items():
        if sheet_name not in workbook.sheetnames:
            print(f"⚠ 工作表 '{sheet_name}' 不存在，跳过")
            continue

        ws = workbook[sheet_name]

        for cell_ref, value in sheet_data.items():
            try:
                # 检查是否是命名范围
                if cell_ref in workbook.defined_names:
                    # 处理命名范围
                    destinations = workbook.defined_names[cell_ref].destinations
                    for title, coord in destinations:
                        ws = workbook[title]
                        ws[coord] = value
                else:
                    # 直接单元格引用
                    ws[cell_ref] = value

                print(f"  ✓ {sheet_name}!{cell_ref} = {value}")
            except Exception as e:
                print(f"  ✗ 填充失败 {sheet_name}!{cell_ref}: {e}")


def fill_template_with_dataframe(template_path, df, sheet_name='Data', start_cell='A1'):
    """
    将 DataFrame 数据填充到模板的指定位置

    参数:
        template_path: str - 模板文件路径
        df: DataFrame - 要填充的数据
        sheet_name: str - 目标工作表名称
        start_cell: str - 起始单元格位置
    """
    wb = load_workbook(template_path)

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    ws = wb[sheet_name]

    # 解析起始单元格
    from openpyxl.utils import column_index_from_string, get_column_letter

    start_col = column_index_from_string(start_cell[0])
    start_row = int(start_cell[1:]) if len(start_cell) > 1 else 1

    # 写入列名
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 写入数据
    for row_idx, row_data in enumerate(df.values, start=start_row + 1):
        for col_idx, value in enumerate(row_data, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)

    print(f"✓ 已将 {len(df)} 行数据填充到 {sheet_name}!{start_cell}")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'filled_report_{timestamp}.xlsx'
    wb.save(output_file)

    return output_file


def create_invoice_from_template(template_path, invoice_data):
    """
    从模板生成发票

    参数:
        template_path: str - 发票模板路径
        invoice_data: dict - 发票数据
            {
                'invoice_number': 'INV-001',
                'date': datetime.now(),
                'customer_name': 'ABC Company',
                'items': [
                    {'description': 'Product A', 'quantity': 10, 'price': 100},
                    {'description': 'Product B', 'quantity': 5, 'price': 200}
                ]
            }
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # 填充表头信息
    ws['B2'] = invoice_data.get('invoice_number', '')
    ws['B3'] = invoice_data.get('date', datetime.now())
    ws['B4'] = invoice_data.get('customer_name', '')

    # 填充明细项
    items = invoice_data.get('items', [])
    start_row = 7  # 假设明细从第 7 行开始

    for idx, item in enumerate(items):
        row = start_row + idx
        ws.cell(row=row, column=1, value=idx + 1)  # 序号
        ws.cell(row=row, column=2, value=item.get('description', ''))  # 描述
        ws.cell(row=row, column=3, value=item.get('quantity', 0))  # 数量
        ws.cell(row=row, column=4, value=item.get('price', 0))  # 单价
        # 小计公式
        ws.cell(row=row, column=5, value=f'=C{row}*D{row}')

    # 总计公式
    total_row = start_row + len(items) + 1
    ws.cell(row=total_row, column=4, value='Total:')
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f'=SUM(E{start_row}:E{start_row + len(items) - 1})')
    ws.cell(row=total_row, column=5).font = Font(bold=True)

    # 保存
    invoice_num = invoice_data.get('invoice_number', 'INV')
    output_file = f'Invoice_{invoice_num}.xlsx'
    wb.save(output_file)

    print(f"✓ 发票已生成: {output_file}")
    return output_file


def batch_fill_templates(template_path, data_list):
    """
    批量生成报告

    参数:
        template_path: str - 模板路径
        data_list: list - 数据字典列表

    返回:
        list - 生成的文件路径列表
    """
    output_files = []

    for idx, data in enumerate(data_list, 1):
        print(f"\n正在处理第 {idx}/{len(data_list)} 个报告...")

        # 生成唯一文件名
        identifier = data.get('identifier', f'report_{idx}')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'{identifier}_{timestamp}.xlsx'

        # 填充模板
        result = fill_template(template_path, data, output_file)
        output_files.append(result)

    print(f"\n✓ 批量生成完成！共生成 {len(output_files)} 个文件")
    return output_files


# 示例用法
if __name__ == "__main__":
    print("Excel 模板填充示例\n")

    # 示例 1: 基本单元格填充
    print("=" * 50)
    print("示例 1: 基本单元格填充")
    print("=" * 50)

    # 注意：需要先创建模板文件
    # template_path = '../templates/business_report.xlsx'

    sample_data = {
        'Sheet1': {
            'B2': 'Annual Sales Report 2024',
            'B4': 'Total Revenue',
            'C4': 1250000,
            'B5': 'Total Orders',
            'C5': 3500,
            'B6': 'Average Order Value',
            'C6': '=C4/C5'  # 公式
        }
    }

    # fill_template(template_path, sample_data, 'filled_report.xlsx')

    # 示例 2: DataFrame 填充
    print("\n" + "=" * 50)
    print("示例 2: DataFrame 数据填充")
    print("=" * 50)

    sales_data = {
        'Product': ['A', 'B', 'C', 'D'],
        'Q1': [1000, 1500, 800, 2000],
        'Q2': [1200, 1400, 900, 2100],
        'Q3': [1100, 1600, 850, 1900],
        'Q4': [1300, 1700, 950, 2200]
    }
    df = pd.DataFrame(sales_data)

    # fill_template_with_dataframe(template_path, df, 'Sales Data', 'A10')

    # 示例 3: 发票生成
    print("\n" + "=" * 50)
    print("示例 3: 发票生成")
    print("=" * 50)

    invoice_data = {
        'invoice_number': 'INV-2024-001',
        'date': datetime.now(),
        'customer_name': 'ABC Corporation',
        'items': [
            {'description': 'Laptop Computer', 'quantity': 5, 'price': 1200},
            {'description': 'Wireless Mouse', 'quantity': 10, 'price': 25},
            {'description': 'USB Cable', 'quantity': 15, 'price': 8}
        ]
    }

    # invoice_template = '../templates/invoice_template.xlsx'
    # create_invoice_from_template(invoice_template, invoice_data)

    # 示例 4: 批量生成
    print("\n" + "=" * 50)
    print("示例 4: 批量生成报告")
    print("=" * 50)

    batch_data = [
        {
            'identifier': 'report_jan',
            'Sheet1': {'B2': 'January Report', 'C4': 100000}
        },
        {
            'identifier': 'report_feb',
            'Sheet1': {'B2': 'February Report', 'C4': 120000}
        },
        {
            'identifier': 'report_mar',
            'Sheet1': {'B2': 'March Report', 'C4': 110000}
        }
    ]

    # batch_fill_templates(template_path, batch_data)

    print("\n" + "=" * 50)
    print("提示：取消注释相关代码并提供模板文件后运行")
    print("=" * 50)
