#!/usr/bin/env python3
"""
高级 Excel 报表生成示例

演示高级功能：图表、条件格式、数据透视表、公式等
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta


def create_advanced_report(df, output_file='advanced_report.xlsx'):
    """
    创建包含高级功能的综合报表

    功能包括:
    - 多工作表
    - 图表可视化
    - 条件格式
    - Excel 表格
    - 公式计算
    """
    print(f"正在生成高级报表: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. 原始数据工作表
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # 2. 数据透视汇总
        create_pivot_summary(df, writer)

        # 3. 趋势分析
        create_trend_analysis(df, writer)

        # 获取 workbook
        workbook = writer.book

        # 4. 添加图表
        add_charts(workbook, df)

        # 5. 应用条件格式
        apply_conditional_formatting(workbook)

        # 6. 创建仪表板
        create_dashboard(workbook, df)

    # 7. 后处理：添加公式
    add_formulas(output_file)

    print(f"✓ 高级报表已生成: {output_file}")
    return output_file


def create_pivot_summary(df, writer):
    """创建数据透视汇总"""
    # 按类别和区域汇总
    if 'Category' in df.columns and 'Region' in df.columns:
        pivot = pd.pivot_table(
            df,
            values='Sales' if 'Sales' in df.columns else df.columns[0],
            index='Category',
            columns='Region',
            aggfunc='sum',
            margins=True,
            margins_name='Total'
        )

        pivot.to_excel(writer, sheet_name='Pivot Summary')

        # 格式化透视表
        ws = writer.sheets['Pivot Summary']
        format_pivot_table(ws)

        print("✓ 已创建数据透视汇总")


def format_pivot_table(worksheet):
    """格式化透视表"""
    # 标题行格式
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # 第一列格式
    for cell in worksheet['A']:
        if cell.row > 1:
            cell.font = Font(bold=True)

    # 合计行格式
    if worksheet.max_row > 1:
        for cell in worksheet[worksheet.max_row]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


def create_trend_analysis(df, writer):
    """创建趋势分析工作表"""
    # 如果有日期列，按日期汇总
    date_columns = df.select_dtypes(include=['datetime64']).columns

    if len(date_columns) > 0:
        date_col = date_columns[0]
        df_trend = df.groupby(pd.Grouper(key=date_col, freq='M')).agg({
            'Sales': ['sum', 'mean', 'count']
        }).reset_index()

        df_trend.to_excel(writer, sheet_name='Trend Analysis', index=False)
        print("✓ 已创建趋势分析")
    else:
        print("⚠ 未找到日期列，跳过趋势分析")


def add_charts(workbook, df):
    """添加各种图表"""
    # 1. 柱状图 - 销售对比
    add_bar_chart(workbook)

    # 2. 折线图 - 趋势
    add_line_chart(workbook)

    # 3. 饼图 - 占比
    add_pie_chart(workbook)

    print("✓ 已添加图表")


def add_bar_chart(workbook):
    """添加柱状图"""
    if 'Pivot Summary' not in workbook.sheetnames:
        return

    ws = workbook['Pivot Summary']

    # 创建柱状图
    chart = BarChart()
    chart.title = "Sales by Category and Region"
    chart.style = 10
    chart.y_axis.title = 'Sales Amount'
    chart.x_axis.title = 'Category'

    # 数据引用（排除合计行）
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row-1, max_col=ws.max_column)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row-1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.shape = 4

    # 添加图表到工作表
    ws.add_chart(chart, "H2")


def add_line_chart(workbook):
    """添加折线图"""
    if 'Trend Analysis' not in workbook.sheetnames:
        return

    ws = workbook['Trend Analysis']

    chart = LineChart()
    chart.title = "Sales Trend Over Time"
    chart.style = 12
    chart.y_axis.title = 'Sales'
    chart.x_axis.title = 'Date'

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)

    ws.add_chart(chart, "F2")


def add_pie_chart(workbook):
    """添加饼图"""
    if 'Pivot Summary' not in workbook.sheetnames:
        return

    ws = workbook['Pivot Summary']

    chart = PieChart()
    chart.title = "Sales Distribution"
    chart.style = 10

    # 使用第一列数据
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row-1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row-1)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    ws.add_chart(chart, "H18")


def apply_conditional_formatting(workbook):
    """应用条件格式"""
    if 'Raw Data' in workbook.sheetnames:
        ws = workbook['Raw Data']

        # 1. 色阶格式 - 销售额
        ws.conditional_formatting.add(
            'B2:B1000',  # 假设销售额在 B 列
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )

        # 2. 数据条 - 可视化数值
        ws.conditional_formatting.add(
            'C2:C1000',  # 假设利润在 C 列
            DataBarRule(
                start_type='min', start_value=0,
                end_type='max', end_value=100,
                color="4472C4"
            )
        )

        # 3. 高亮低于平均值的单元格
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws.conditional_formatting.add(
            'B2:B1000',
            CellIsRule(operator='lessThan', formula=['AVERAGE($B$2:$B$1000)'], fill=red_fill)
        )

        print("✓ 已应用条件格式")


def create_dashboard(workbook, df):
    """创建仪表板工作表"""
    # 创建新工作表
    dashboard = workbook.create_sheet('Dashboard', 0)

    # 设置标题
    dashboard['A1'] = 'Business Performance Dashboard'
    dashboard['A1'].font = Font(size=16, bold=True, color='1F4E78')
    dashboard.merge_cells('A1:F1')
    dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
    dashboard.row_dimensions[1].height = 30

    # KPI 指标
    kpis = {
        'Total Sales': df['Sales'].sum() if 'Sales' in df.columns else 0,
        'Average Sales': df['Sales'].mean() if 'Sales' in df.columns else 0,
        'Total Orders': len(df),
        'Unique Products': df['Product'].nunique() if 'Product' in df.columns else 0
    }

    row = 3
    for kpi_name, kpi_value in kpis.items():
        # KPI 名称
        dashboard.cell(row=row, column=1, value=kpi_name)
        dashboard.cell(row=row, column=1).font = Font(bold=True, size=12)

        # KPI 值
        dashboard.cell(row=row, column=2, value=kpi_value)
        dashboard.cell(row=row, column=2).font = Font(size=14, color='0070C0')
        dashboard.cell(row=row, column=2).number_format = '#,##0.00'

        row += 1

    # 调整列宽
    dashboard.column_dimensions['A'].width = 20
    dashboard.column_dimensions['B'].width = 15

    print("✓ 已创建仪表板")


def add_formulas(filename):
    """添加 Excel 公式"""
    wb = load_workbook(filename)

    if 'Raw Data' in wb.sheetnames:
        ws = wb['Raw Data']

        # 在最后一行添加合计公式
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1, value='TOTAL')
        ws.cell(row=last_row, column=1).font = Font(bold=True, size=12)

        # 销售额合计
        ws.cell(row=last_row, column=2, value=f'=SUM(B2:B{last_row-2})')
        ws.cell(row=last_row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=last_row, column=2).number_format = '#,##0.00'

        # 利润合计
        if ws.max_column >= 3:
            ws.cell(row=last_row, column=3, value=f'=SUM(C2:C{last_row-2})')
            ws.cell(row=last_row, column=3).font = Font(bold=True, size=12)
            ws.cell(row=last_row, column=3).number_format = '#,##0.00'

    wb.save(filename)
    print("✓ 已添加公式")


# 示例用法
if __name__ == "__main__":
    # 生成示例数据
    np.random.seed(42)
    dates = pd.date_range('2024-01-01', periods=100, freq='D')

    sample_data = {
        'Date': dates,
        'Product': np.random.choice(['Product A', 'Product B', 'Product C', 'Product D'], 100),
        'Category': np.random.choice(['Electronics', 'Clothing', 'Food', 'Books'], 100),
        'Region': np.random.choice(['North', 'South', 'East', 'West'], 100),
        'Sales': np.random.randint(100, 2000, 100),
        'Profit': np.random.randint(20, 500, 100),
        'Quantity': np.random.randint(1, 50, 100)
    }

    df = pd.DataFrame(sample_data)

    # 生成高级报表
    create_advanced_report(df, 'advanced_sample_report.xlsx')

    print("\n高级报表生成完成！")
    print("查看文件: advanced_sample_report.xlsx")
